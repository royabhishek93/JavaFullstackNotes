#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Data to merge (Batch 3: Seq 21-30)
new_properties = [
    {
        "Seq": 21, "Project": "Sumadhura Capitol Residences", "Area": "Whitefield",
        "Builder": "Sumadhura", "Builder Grade": "A", "Builder Score": 9.5,
        "Price Min (₹/sq.ft)": 12500, "Price Max (₹/sq.ft)": 14500,
        "Nearby IT Hubs": "ITPL, EPIP Zone, GR Tech Park",
        "IT Distance": "1–3 km", "SAP Labs Distance": "2–3 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "2 km",
        "Society Size": "Medium (≈405 units)", "Possession Status": "Under Construction",
        "Rental Demand": 5, "Resale Potential": 5, "Overall Rating": 9.4,
        "Key Notes": "Premium project beside ITPL"
    },
    {
        "Seq": 22, "Project": "Prestige Jade Pavilion", "Area": "Marathahalli / ECC Road",
        "Builder": "Prestige", "Builder Grade": "A+", "Builder Score": 9.9,
        "Price Min (₹/sq.ft)": 13500, "Price Max (₹/sq.ft)": 16000,
        "Nearby IT Hubs": "ITPL, EPIP Zone, Bagmane Tech Park",
        "IT Distance": "3–6 km", "SAP Labs Distance": "5–6 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Ready",
        "Rental Demand": 5, "Resale Potential": 5, "Overall Rating": 9.5,
        "Key Notes": "Premium Prestige development"
    },
    {
        "Seq": 23, "Project": "TVS Emerald (Project)", "Area": "Thanisandra / South Bengaluru*",
        "Builder": "TVS Emerald", "Builder Grade": "A", "Builder Score": 9.2,
        "Price Min (₹/sq.ft)": 11000, "Price Max (₹/sq.ft)": 13000,
        "Nearby IT Hubs": "Manyata Tech Park / Electronic City*",
        "IT Distance": "4–8 km", "SAP Labs Distance": "N/A",
        "Nearest Metro": "Depends on project", "Metro Distance": "3–5 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 9.1,
        "Key Notes": "Specify exact project name (TVS Emerald has multiple projects)"
    },
    {
        "Seq": 24, "Project": "Nambiar District 25", "Area": "Muthanallur Road",
        "Builder": "Nambiar Builders", "Builder Grade": "A", "Builder Score": 9.3,
        "Price Min (₹/sq.ft)": 10500, "Price Max (₹/sq.ft)": 12500,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ, RMZ Ecoworld",
        "IT Distance": "6–9 km", "SAP Labs Distance": "10–12 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "5 km",
        "Society Size": "Large Township", "Possession Status": "Under Construction",
        "Rental Demand": 5, "Resale Potential": 5, "Overall Rating": 9.3,
        "Key Notes": "Premium township with strong appreciation potential"
    },
    {
        "Seq": 25, "Project": "Sterling Ascentia", "Area": "Whitefield",
        "Builder": "Sterling Developers", "Builder Grade": "A-", "Builder Score": 8.8,
        "Price Min (₹/sq.ft)": 10500, "Price Max (₹/sq.ft)": 12000,
        "Nearby IT Hubs": "ITPL, EPIP Zone",
        "IT Distance": "3–5 km", "SAP Labs Distance": "4–5 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "3 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 8.9,
        "Key Notes": "Boutique premium community"
    },
    {
        "Seq": 26, "Project": "Orchid Lakeview", "Area": "Whitefield",
        "Builder": "Goyal & Co. / Orchid*", "Builder Grade": "A-", "Builder Score": 8.8,
        "Price Min (₹/sq.ft)": 10000, "Price Max (₹/sq.ft)": 12000,
        "Nearby IT Hubs": "ITPL, EPIP Zone",
        "IT Distance": "3–5 km", "SAP Labs Distance": "4–5 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "3 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 8.9,
        "Key Notes": "Verify exact developer/project variant"
    },
    {
        "Seq": 27, "Project": "Klassik Landmark", "Area": "Sarjapur Road",
        "Builder": "Klassik Enterprises", "Builder Grade": "A-", "Builder Score": 8.8,
        "Price Min (₹/sq.ft)": 9800, "Price Max (₹/sq.ft)": 11500,
        "Nearby IT Hubs": "RMZ Ecoworld, Embassy Tech Village, Wipro SEZ",
        "IT Distance": "2–5 km", "SAP Labs Distance": "5–6 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Ready",
        "Rental Demand": 5, "Resale Potential": 4, "Overall Rating": 9.0,
        "Key Notes": "Excellent rental location"
    },
    {
        "Seq": 28, "Project": "Ardente Pine Grove", "Area": "Hennur Road",
        "Builder": "Ardente Realty", "Builder Grade": "A-", "Builder Score": 8.7,
        "Price Min (₹/sq.ft)": 9500, "Price Max (₹/sq.ft)": 11000,
        "Nearby IT Hubs": "Manyata Tech Park",
        "IT Distance": "5–7 km", "SAP Labs Distance": "N/A",
        "Nearest Metro": "Nagawara", "Metro Distance": "5 km",
        "Society Size": "Small–Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 8.8,
        "Key Notes": "Boutique Hennur project (Scribd)"
    },
    {
        "Seq": 29, "Project": "Jana Jeeva Splendour", "Area": "Sarjapur Road",
        "Builder": "Jana Jeeva Estates", "Builder Grade": "B+", "Builder Score": 8.3,
        "Price Min (₹/sq.ft)": 8800, "Price Max (₹/sq.ft)": 10000,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ",
        "IT Distance": "5–8 km", "SAP Labs Distance": "9–10 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "5 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 3, "Overall Rating": 8.5,
        "Key Notes": "Budget option near IT corridor"
    },
    {
        "Seq": 30, "Project": "Brigade Gem", "Area": "Sarjapur Road",
        "Builder": "Brigade Group", "Builder Grade": "A+", "Builder Score": 9.8,
        "Price Min (₹/sq.ft)": 12500, "Price Max (₹/sq.ft)": 14500,
        "Nearby IT Hubs": "RMZ Ecoworld, Embassy Tech Village, Wipro SEZ",
        "IT Distance": "3–6 km", "SAP Labs Distance": "6–7 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 5, "Resale Potential": 5, "Overall Rating": 9.4,
        "Key Notes": "Brigade quality, premium specifications"
    }
]

# Load existing workbook
file_path = "/Users/I771246/Abhi Personal/JavaFullstackNotes/DREAM_HOME/Bangalore_Property_Comparison_Analysis.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Get existing headers from row 1
headers = [cell.value for cell in ws[1]]
print(f"Existing headers: {headers[:19]}")  # Print first 19 relevant headers

# Find the last row with data
last_row = ws.max_row
print(f"Current last row: {last_row}")

# Define styles for consistency
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Add new properties
for prop in new_properties:
    last_row += 1
    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue

        cell = ws.cell(row=last_row, column=col_idx)
        cell.value = prop.get(header, "")
        cell.border = thin_border

        # Apply alignment based on column type
        if header in ["Project", "Area", "Builder", "Nearby IT Hubs", "Key Notes"]:
            cell.alignment = left_alignment
        else:
            cell.alignment = center_alignment

        # Bold and highlight for A+ builders
        if header == "Builder Grade" and prop.get("Builder Grade") == "A+":
            cell.font = Font(bold=True)
        elif header == "Builder Grade" and prop.get("Builder Grade") == "A":
            cell.font = Font(bold=True)

        # Bold high builder scores
        if header == "Builder Score" and prop.get("Builder Score", 0) >= 9.5:
            cell.font = Font(bold=True)
        elif header == "Builder Score" and prop.get("Builder Score", 0) >= 9.0:
            cell.font = Font(bold=True)

        # Highlight high overall ratings
        if header == "Overall Rating" and prop.get("Overall Rating", 0) >= 9.0:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

# Save the workbook
wb.save(file_path)
print(f"\nSuccessfully added {len(new_properties)} properties to the Excel file!")
print(f"Total rows now: {ws.max_row}")
print(f"New properties added: Seq {new_properties[0]['Seq']} to {new_properties[-1]['Seq']}")
