#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Data to merge (Batch 4: Seq 31-39)
new_properties = [
    {
        "Seq": 31, "Project": "Big Banyan Roots", "Area": "Mysore Road",
        "Builder": "Big Banyan", "Builder Grade": "A-", "Builder Score": 8.7,
        "Price Min (₹/sq.ft)": 8500, "Price Max (₹/sq.ft)": 10000,
        "Nearby IT Hubs": "Global Village Tech Park, Kalyani Tech Park",
        "IT Distance": "4–7 km", "SAP Labs Distance": "20+ km",
        "Nearest Metro": "Challaghatta", "Metro Distance": "6 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 8.8,
        "Key Notes": "Good value in West Bangalore"
    },
    {
        "Seq": 32, "Project": "Abhee Silicon Shine", "Area": "Gunjur",
        "Builder": "Abhee Ventures", "Builder Grade": "A-", "Builder Score": 8.8,
        "Price Min (₹/sq.ft)": 10000, "Price Max (₹/sq.ft)": 11800,
        "Nearby IT Hubs": "RMZ Ecoworld, Embassy Tech Village, RGA Tech Park",
        "IT Distance": "5–7 km", "SAP Labs Distance": "7–8 km",
        "Nearest Metro": "Kadubeesanahalli (Future)", "Metro Distance": "5 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 5, "Resale Potential": 4, "Overall Rating": 9.0,
        "Key Notes": "Strong Abhee project near ORR"
    },
    {
        "Seq": 33, "Project": "Siroya Environ", "Area": "Whitefield",
        "Builder": "Siroya Developers", "Builder Grade": "B+", "Builder Score": 8.4,
        "Price Min (₹/sq.ft)": 9500, "Price Max (₹/sq.ft)": 11000,
        "Nearby IT Hubs": "ITPL, EPIP Zone, GR Tech Park",
        "IT Distance": "3–5 km", "SAP Labs Distance": "4–5 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "3 km",
        "Society Size": "Small", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 3, "Overall Rating": 8.6,
        "Key Notes": "Boutique community"
    },
    {
        "Seq": 34, "Project": "Solcrest by Bricks & Milestones", "Area": "Budigere Cross",
        "Builder": "Bricks & Milestones", "Builder Grade": "A-", "Builder Score": 8.9,
        "Price Min (₹/sq.ft)": 10500, "Price Max (₹/sq.ft)": 12500,
        "Nearby IT Hubs": "ITPL, Bagmane Tech Park, Brigade Tech Park",
        "IT Distance": "8–12 km", "SAP Labs Distance": "10–12 km",
        "Nearest Metro": "KR Puram", "Metro Distance": "8 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 9.0,
        "Key Notes": "Strong boutique builder"
    },
    {
        "Seq": 35, "Project": "Vaishnavi Serene", "Area": "Yelahanka",
        "Builder": "Vaishnavi Group", "Builder Grade": "A", "Builder Score": 9.2,
        "Price Min (₹/sq.ft)": 10500, "Price Max (₹/sq.ft)": 12500,
        "Nearby IT Hubs": "Manyata Tech Park, Kirloskar Tech Park",
        "IT Distance": "8–10 km", "SAP Labs Distance": "15+ km",
        "Nearest Metro": "Yelahanka (Future)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 5, "Overall Rating": 9.1,
        "Key Notes": "Premium Vaishnavi quality"
    },
    {
        "Seq": 36, "Project": "Sobha Dewflower", "Area": "Hennur Road",
        "Builder": "Sobha", "Builder Grade": "A+", "Builder Score": 9.9,
        "Price Min (₹/sq.ft)": 13000, "Price Max (₹/sq.ft)": 15500,
        "Nearby IT Hubs": "Manyata Tech Park, Kirloskar Tech Park",
        "IT Distance": "5–7 km", "SAP Labs Distance": "15+ km",
        "Nearest Metro": "Nagawara", "Metro Distance": "5 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 5, "Resale Potential": 5, "Overall Rating": 9.4,
        "Key Notes": "Premium Sobha specifications"
    },
    {
        "Seq": 37, "Project": "Godrej Eternity", "Area": "Kanakapura Road",
        "Builder": "Godrej Properties", "Builder Grade": "A", "Builder Score": 9.5,
        "Price Min (₹/sq.ft)": 10500, "Price Max (₹/sq.ft)": 12500,
        "Nearby IT Hubs": "Global Village Tech Park, Kalyani Tech Park",
        "IT Distance": "8–12 km", "SAP Labs Distance": "22+ km",
        "Nearest Metro": "Vajarahalli", "Metro Distance": "2 km",
        "Society Size": "Large", "Possession Status": "Ready",
        "Rental Demand": 4, "Resale Potential": 5, "Overall Rating": 9.2,
        "Key Notes": "Excellent metro connectivity"
    },
    {
        "Seq": 38, "Project": "Brigade Woods", "Area": "Whitefield",
        "Builder": "Brigade", "Builder Grade": "A+", "Builder Score": 9.8,
        "Price Min (₹/sq.ft)": 11500, "Price Max (₹/sq.ft)": 13500,
        "Nearby IT Hubs": "ITPL, EPIP Zone, GR Tech Park",
        "IT Distance": "2–4 km", "SAP Labs Distance": "3–4 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "3 km",
        "Society Size": "Medium", "Possession Status": "Ready",
        "Rental Demand": 5, "Resale Potential": 5, "Overall Rating": 9.3,
        "Key Notes": "Excellent Whitefield location"
    },
    {
        "Seq": 39, "Project": "Brigade Orchards", "Area": "Devanahalli",
        "Builder": "Brigade", "Builder Grade": "A+", "Builder Score": 9.8,
        "Price Min (₹/sq.ft)": 9500, "Price Max (₹/sq.ft)": 11500,
        "Nearby IT Hubs": "KIADB Aerospace Park, IFCI Financial City, Manyata Tech Park",
        "IT Distance": "8–12 km (Aerospace Park), 25–30 km (Manyata)", "SAP Labs Distance": "35+ km",
        "Nearest Metro": "Airport Metro (Upcoming)", "Metro Distance": "6 km",
        "Society Size": "Mega Township (130–135 acres, multiple phases)", "Possession Status": "Mixed (Ready + Under Construction)",
        "Rental Demand": 4, "Resale Potential": 5, "Overall Rating": 9.1,
        "Key Notes": "Integrated township with school, sports arena, offices and retail; not a boutique community. (Brigade Orchards)"
    }
]

# Load existing workbook
file_path = "/Users/I771246/Abhi Personal/JavaFullstackNotes/DREAM_HOME/Bangalore_Property_Comparison_Analysis.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Get existing headers from row 1
headers = [cell.value for cell in ws[1]]
print(f"Existing headers: {headers[:19]}")  # Print first 19 relevant headers

# Find the last row with data
last_row = ws.max_row
print(f"Current last row: {last_row}")

# Define styles for consistency
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Add new properties
for prop in new_properties:
    last_row += 1
    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue

        cell = ws.cell(row=last_row, column=col_idx)
        cell.value = prop.get(header, "")
        cell.border = thin_border

        # Apply alignment based on column type
        if header in ["Project", "Area", "Builder", "Nearby IT Hubs", "Key Notes"]:
            cell.alignment = left_alignment
        else:
            cell.alignment = center_alignment

        # Bold and highlight for A+ builders
        if header == "Builder Grade" and prop.get("Builder Grade") == "A+":
            cell.font = Font(bold=True)
        elif header == "Builder Grade" and prop.get("Builder Grade") == "A":
            cell.font = Font(bold=True)

        # Bold high builder scores
        if header == "Builder Score" and prop.get("Builder Score", 0) >= 9.5:
            cell.font = Font(bold=True)
        elif header == "Builder Score" and prop.get("Builder Score", 0) >= 9.0:
            cell.font = Font(bold=True)

        # Highlight high overall ratings
        if header == "Overall Rating" and prop.get("Overall Rating", 0) >= 9.0:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

# Save the workbook
wb.save(file_path)
print(f"\nSuccessfully added {len(new_properties)} properties to the Excel file!")
print(f"Total rows now: {ws.max_row}")
print(f"New properties added: Seq {new_properties[0]['Seq']} to {new_properties[-1]['Seq']}")
