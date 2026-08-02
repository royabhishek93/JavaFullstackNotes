#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Load existing workbook
wb = openpyxl.load_workbook('Bangalore_Property_Comparison_Analysis.xlsx')

# Get or create sheets for different property lists
def get_or_create_sheet(name):
    if name in wb.sheetnames:
        return wb[name]
    return wb.create_sheet(name)

# Property data from your tables
properties_data = {
    "Low Rise Projects": [
        ["Project", "Area", "Floors", "Approx. Units", "Builder"],
        ["Brigade Gem", "Sarjapur Road", "B+G+4", 138, "Brigade"],
        ["Adarsh Lumina", "Gattahalli", "B+G+4", 440, "Adarsh"],
        ["Abhee Silicon Shine", "Mullur", "2B+G+4", "276-462", "Abhee"],
        ["Purva Whitehall", "Whitefield", "Low-rise", "Boutique", "Puravankara"],
        ["Brigade Woods", "Whitefield", "Low-rise", "Medium", "Brigade"],
        ["Big Banyan Roots", "Sarjapur Road", "Low-rise", "Boutique", "Big Banyan"],
        ["Surbacon Cedar", "Electronic City", "G+4/G+5", 156, "Surbacon"],
        ["Brigade Orchards - Laurel & Maple", "Devanahalli", "Low-rise", "Medium", "Brigade"],
        ["Godrej Eternity", "Kanakapura Road", "Low-rise", "Medium", "Godrej"],
        ["Vaishnavi Serene", "Yelahanka", "Low-rise", "Medium", "Vaishnavi"],
        ["Isha Misty Green", "Whitefield", "Low-rise", "Medium", "Isha"],
        ["Sobha Dewflower", "JP Nagar", "Low-rise", "Boutique", "Sobha"],
    ],

    "Status Overview": [
        ["Project", "Floors", "Approx. Units", "Builder", "Status"],
        ["Brigade Gem", "B+G+4", 138, "Brigade", "Ready"],
        ["Big Banyan Roots", "B+G+4", "Boutique", "Big Banyan", "Ready"],
        ["Adarsh Lumina", "B+G+4", 440, "Adarsh", "Under construction"],
        ["Abhee Silicon Shine", "2B+G+4", 462, "Abhee", "Under construction"],
        ["Purva Whitehall", "Low-rise", "Boutique", "Puravankara", "Ready"],
        ["Surbacon Cedar", "G+4/G+5", 156, "Surbacon", "Near possession"],
        ["Bhavya Pristine", "G+5", 100, "Bhavya", "Ready"],
        ["MH Tirumala Heights", "G+5", 84, "MH Promoters", "Ready"],
    ],

    "Top Rated Projects": [
        ["Project", "Floors", "Approx. Units", "Builder", "My Rating"],
        ["Brigade Gem", "B+G+4", 138, "Brigade", "⭐ 9.3"],
        ["Adarsh Lumina", "B+G+4", 440, "Adarsh", "⭐ 9.3"],
        ["Big Banyan Roots", "Low-rise", "Boutique", "Big Banyan", "⭐ 9.0"],
        ["Purva Whitehall", "Low-rise", "Boutique", "Puravankara", "⭐ 9.0"],
        ["Abhee Silicon Shine", "2B+G+4", "276-462", "Abhee", "⭐ 9.0"],
        ["Surbacon Cedar", "G+4/G+5", 156, "Surbacon", "⭐ 8.8"],
        ["Bhavya Pristine", "G+5", 100, "Bhavya", "⭐ 8.9"],
        ["Poorvi Enchanting", "G+11", 99, "Poorvi", "⭐ 9.1"],
        ["MH Tirumala Heights", "G+5", 84, "MH Promoters", "⭐ 8.8"],
    ],

    "Small Societies (<200 Units)": [
        ["Project", "Approx. Units", "Status", "Builder", "Meets <200?"],
        ["Brigade Gem", 138, "Ready", "Brigade", "✅"],
        ["Abhee Silicon Shine", "~180", "Under construction", "Abhee", "✅"],
        ["DSR The Courtyard", "~120", "Ready", "DSR", "✅"],
        ["Subha Esperanza", "<200", "Under construction", "Subha", "✅"],
        ["Wone Royal Nest", 104, "Under construction", "Wone", "✅"],
        ["Bhavya Pristine", 100, "Ready", "Bhavya", "✅"],
        ["Poorvi Enchanting", 99, "Ready", "Poorvi", "✅"],
        ["Surbacon Cedar", 156, "Near possession", "Surbacon", "✅"],
        ["SLS Splendor", "~180", "Near possession", "SLS", "✅"],
        ["SLS Sunflower", "~140", "Near possession", "SLS", "✅"],
        ["TG Ascent", "<200", "Under construction", "TG Developers", "✅"],
    ],

    "Boutique Projects": [
        ["Project", "Approx. Units", "Status", "Builder Grade", "My Rating"],
        ["Poorvi Enchanting", 99, "Ready", "B+", "⭐ 9.1"],
        ["Bhavya Pristine", 100, "Ready", "B+", "⭐ 8.9"],
        ["Surbacon Cedar", 156, "Near Possession", "B+", "⭐ 8.8"],
        ["SLS Splendor", "~180", "Near Possession", "B", "⭐ 9.0"],
        ["SLS Sunflower", "~140", "Near Possession", "B", "⭐ 8.9"],
        ["TG Ascent", "<200", "Under Construction", "B+", "⭐ 9.0"],
        ["Ahad Excellencia", "~120", "Ready", "B+", "⭐ 8.9"],
        ["Keerthi Regalia", "<200", "Ready", "B", "⭐ 8.8"],
        ["SUYUG The 1", "<200", "Under Construction", "B+", "⭐ 9.0"],
        ["Wone Royal Nest", 104, "Under Construction", "B", "⭐ 8.7"],
        ["Subha Esperanza", "<200", "Under Construction", "B+", "⭐ 9.0"],
        ["GR Sitara", "<200", "Under Construction", "B", "⭐ 8.8"],
        ["Narmada Shri Heights", "<200", "Ready", "B", "⭐ 8.7"],
        ["Sri Balaji Vaibhav", "<200", "Ready", "B", "⭐ 8.7"],
        ["MH Tirumala Heights", 84, "Ready", "B", "⭐ 8.8"],
    ],

    "Premium Comparison": [
        ["Project", "Units", "My Rating"],
        ["Sumadhura Capitol Residences", "500+", "⭐ 9.5"],
        ["DSR Parkway", "~220", "⭐ 9.4"],
        ["Adarsh Lumina", 440, "⭐ 9.3"],
        ["Sterling Ascentia", "~392", "⭐ 9.4"],
        ["Adarsh Tropica", "400+", "⭐ 9.1"],
        ["Poorvi Enchanting", 99, "⭐ 9.1"],
        ["Orchid Lakeview", 336, "⭐ 9.2"],
        ["Prestige Jade Pavilion", "~470", "⭐ 9.2"],
        ["Emmanuel Heights", "~290", "⭐ 9.0"],
    ],

    "Detailed Ratings - Top Picks": [
        ["Project", "Builder", "Approx. Units", "Category", "Possession", "Overall"],
        ["Poorvi Enchanting", "Poorvi", 99, "Boutique", "Ready/Near", "⭐ 9.1"],
        ["Bhavya Pristine", "Bhavya", 100, "Boutique", "Ready", "⭐ 8.9"],
        ["MH Tirumala Heights", "MH Promoters", 84, "Boutique", "Ready", "⭐ 8.8"],
        ["Surbacon Cedar", "Surbacon", 156, "Premium", "Near possession", "⭐ 8.8"],
        ["SLS Sunflower", "SLS", "~140", "Premium", "Near possession", "⭐ 8.9"],
        ["SLS Splendor", "SLS", "~180", "Premium", "Near possession", "⭐ 9.0"],
        ["TG Ascent", "TG Developers", "<200", "Premium", "Under construction", "⭐ 9.0"],
        ["Subha Esperanza", "Subha", "<200", "Premium", "Dec 2026", "⭐ 9.0"],
        ["Keerthi Regalia", "Keerthi Estates", "<200", "Premium", "Ready", "⭐ 8.9"],
        ["SUYUG The 1", "Suyug", "<200", "Boutique Premium", "Under construction", "⭐ 9.0"],
        ["Jana Jeeva Splendour", "Jana Jeeva", "~50", "Boutique", "Ready", "⭐ 8.7"],
    ],

    "Large Projects Comparison": [
        ["Project", "My Rating"],
        ["Sumadhura Capitol Residences", "⭐ 9.5"],
        ["DSR Parkway", "⭐ 9.4"],
        ["TVS Emerald", "⭐ 9.3"],
        ["Prestige Jade Pavilion", "⭐ 9.2"],
        ["Adarsh Tropica", "⭐ 9.1"],
        ["Mana Dale", "⭐ 9.1"],
        ["MH Tirumala Heights", "⭐ 8.8"],
        ["Mahaveer Ranches", "⭐ 8.6"],
    ],
}

# Additional scoring sheets
scoring_data = {
    "Adarsh Lumina - Detailed Score": [
        ["Parameter", "Score"],
        ["Builder Reputation", "⭐ 9.6"],
        ["Location", "⭐ 9.4"],
        ["Construction Quality", "⭐ 9.4"],
        ["IT Connectivity", "⭐ 9.5"],
        ["Rental Demand", "⭐ 9.4"],
        ["Appreciation", "⭐ 9.3"],
        ["Society Size", "⭐ 7.5"],
        ["Overall", "9.2/10"],
    ],

    "Sterling Ascentia - Score": [
        ["Parameter", "Rating"],
        ["Builder Reputation", "⭐ 9.5"],
        ["Location", "⭐ 9.8"],
        ["Construction Quality", "⭐ 9.4"],
        ["IT Connectivity", "⭐ 10.0"],
        ["Rental Demand", "⭐ 9.8"],
        ["Appreciation", "⭐ 9.5"],
        ["Society Size", "⭐ 7.5"],
        ["Overall", "9.4"],
    ],

    "Orchid Lakeview - Score": [
        ["Parameter", "Score"],
        ["Builder Reputation", "⭐ 9.0"],
        ["Location", "⭐ 9.8"],
        ["Construction Quality", "⭐ 9.0"],
        ["IT Connectivity", "⭐ 10.0"],
        ["Rental Demand", "⭐ 9.8"],
        ["Appreciation Potential", "⭐ 9.3"],
        ["Society Size", "⭐ 8.0"],
        ["Overall", "9.2"],
    ],

    "Emmanuel Heights - Score": [
        ["Parameter", "Rating"],
        ["Location", "⭐ 9.2"],
        ["Builder", "⭐ 8.7"],
        ["Construction Quality", "⭐ 9.0"],
        ["Society Size", "⭐ 8.5"],
        ["IT Connectivity", "⭐ 9.3"],
        ["Rental Demand", "⭐ 9.1"],
        ["Appreciation", "⭐ 9.0"],
        ["Overall", "9.0"],
    ],
}

def format_sheet(ws, data):
    """Format a worksheet with data and styling"""
    # Clear existing content if any
    ws.delete_rows(1, ws.max_row)

    # Write data
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Format header row
            if row_idx == 1:
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = ws['A2']

# Create or update all sheets with property data
for sheet_name, data in properties_data.items():
    ws = get_or_create_sheet(sheet_name)
    format_sheet(ws, data)
    print(f"✅ Created/Updated: {sheet_name}")

# Create or update scoring sheets
for sheet_name, data in scoring_data.items():
    ws = get_or_create_sheet(sheet_name)
    format_sheet(ws, data)
    print(f"✅ Created/Updated: {sheet_name}")

# Save the workbook
wb.save('Bangalore_Property_Comparison_Analysis.xlsx')
print(f"\n✨ Excel file updated successfully!")
print(f"📊 Total sheets: {len(wb.sheetnames)}")
print(f"📝 Sheets: {', '.join(wb.sheetnames)}")
