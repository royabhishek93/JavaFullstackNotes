#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Load existing workbook
wb = openpyxl.load_workbook('Bangalore_Property_Comparison_Analysis.xlsx')

# Builder grades data
builder_grades = [
    ["Builder", "Grade", "Confidence", "Notes"],
    ["Prestige", "A+", "⭐⭐⭐⭐⭐", "Tier 1 - Premium"],
    ["Sobha", "A+", "⭐⭐⭐⭐⭐", "Tier 1 - Premium"],
    ["Brigade", "A+", "⭐⭐⭐⭐⭐", "Tier 1 - Premium"],
    ["Adarsh", "A+", "⭐⭐⭐⭐⭐", "Tier 1 - Premium"],
    ["Embassy", "A+", "⭐⭐⭐⭐⭐", "Tier 1 - Premium"],
    ["Godrej Properties", "A+", "⭐⭐⭐⭐⭐", "Tier 1 - Premium"],
    ["Assetz", "A+", "⭐⭐⭐⭐⭐", "Tier 1 - Premium"],
    ["Mahindra Lifespaces", "A+", "⭐⭐⭐⭐⭐", "Tier 1 - Premium"],
    ["Sumadhura", "A+", "⭐⭐⭐⭐⭐", "Tier 1 - Premium"],
    ["Puravankara", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Vaishnavi Group", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Concorde", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Arvind SmartSpaces", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["TVS Emerald", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Century Real Estate", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Keya Homes", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Spectra", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Alembic", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Kolte-Patil", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Phoenix Mills", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Goyal & Co.", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Candeur", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Navami", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Kumar Properties", "A-", "⭐⭐⭐⭐", "Tier 2 - Reliable"],
    ["Mana Projects", "A-", "⭐⭐⭐⭐", "Tier 2 - Reliable"],
    ["Paras Buildtech", "A-", "⭐⭐⭐⭐", "Tier 2 - Reliable"],
    ["Shriram Properties", "A-", "⭐⭐⭐⭐", "Tier 2 - Reliable"],
    ["Casagrand", "A-", "⭐⭐⭐⭐", "Tier 2 - Reliable"],
    ["Trifecta", "A-", "⭐⭐⭐⭐", "Tier 2 - Reliable"],
    ["Suyug", "A-", "⭐⭐⭐⭐", "Tier 2 - Reliable"],
    ["Gopalan", "A-", "⭐⭐⭐⭐", "Tier 2 - Reliable"],
    ["Meenakshi Group", "A-", "⭐⭐⭐⭐", "Tier 2 - Reliable"],
    ["Trinity", "A-", "⭐⭐⭐", "Tier 2 - Reliable"],
    ["Abhee Ventures", "A-", "⭐⭐⭐", "Tier 2 - Reliable"],
    ["Mythri Builders", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Habitat Ventures", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Confident Group", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["CoEvolve Estates", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Sipani Properties", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Jeevan Developers", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Binary Realty", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["SBR Group", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Poorvi", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Bhavya", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Surbacon", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["TG Developers", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Ahad", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Subha", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
    ["Evershine", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Pranvi", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Formist", "B", "⭐⭐☆", "Too new for higher grading"],
    ["Regency Group", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Elegant Builders", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Classique", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Saritha Developers", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Myhna Properties", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["CKPC Properties", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["United Developers", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Amogaya", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Svasa Homes", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Ashed Properties", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["SLS", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["MH Promoters", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Keerthi Estates", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Wone", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["GR", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Jana Jeeva", "B", "⭐⭐☆", "Tier 3 - Budget"],
    ["Big Banyan", "B+", "⭐⭐⭐", "Boutique - Good quality"],
    ["DSR", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Sterling", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Orchid (Goyal & Hariyana)", "A", "⭐⭐⭐⭐☆", "Tier 1 - Solid"],
    ["Emmanuel", "B+", "⭐⭐⭐", "Tier 3 - Mid-tier"],
]

# Create builder grade mapping
builder_map = {}
for row in builder_grades[1:]:  # Skip header
    builder_map[row[0]] = {"grade": row[1], "confidence": row[2], "notes": row[3]}

def format_sheet(ws, data):
    """Format a worksheet with data and styling"""
    # Clear existing content
    ws.delete_rows(1, ws.max_row)

    # Write data
    for row_idx, row_data in enumerate(data, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            # Format header row
            if row_idx == 1:
                cell.font = Font(bold=True, size=11, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column].width = adjusted_width

    # Freeze header row
    ws.freeze_panes = ws['A2']

# Update or create Builder Ratings sheet
if "Builder Ratings" in wb.sheetnames:
    ws = wb["Builder Ratings"]
else:
    ws = wb.create_sheet("Builder Ratings", 1)  # Insert as second sheet

format_sheet(ws, builder_grades)
print(f"✅ Updated: Builder Ratings (with {len(builder_grades)-1} builders)")

# Now update existing project sheets with correct builder grades
sheets_to_update = {
    "Boutique Projects": 3,  # Builder Grade column index
    "Detailed Ratings - Top Picks": 1,  # Builder column index
}

for sheet_name, builder_col_idx in sheets_to_update.items():
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        updated_count = 0

        # Update builder grades based on builder name
        for row_idx in range(2, ws.max_row + 1):  # Skip header
            builder_name = ws.cell(row=row_idx, column=builder_col_idx).value

            if builder_name in builder_map:
                # Find and update Builder Grade column
                for col_idx in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col_idx).value
                    if header == "Builder Grade":
                        ws.cell(row=row_idx, column=col_idx).value = builder_map[builder_name]["grade"]
                        updated_count += 1
                        break

        print(f"✅ Updated {updated_count} builder grades in: {sheet_name}")

# Save the workbook
wb.save('Bangalore_Property_Comparison_Analysis.xlsx')
print(f"\n✨ Excel file updated with verified builder grades!")
print(f"📊 Total builders cataloged: {len(builder_grades)-1}")
