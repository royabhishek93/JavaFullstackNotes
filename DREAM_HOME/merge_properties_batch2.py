#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Data to merge (Batch 2: Seq 11-20)
new_properties = [
    {
        "Seq": 11, "Project": "Ahad Excellencia", "Area": "Choodasandra / Sarjapur Road",
        "Builder": "Ahad Builders", "Builder Grade": "A-", "Builder Score": 8.9,
        "Price Min (₹/sq.ft)": 9000, "Price Max (₹/sq.ft)": 11000,
        "Nearby IT Hubs": "Wipro SEZ, RGA Tech Park, RMZ Ecoworld",
        "IT Distance": "1–4 km", "SAP Labs Distance": "8–10 km",
        "Nearest Metro": "Bommasandra", "Metro Distance": "7 km",
        "Society Size": "Large (505 units)", "Possession Status": "Ready",
        "Rental Demand": 5, "Resale Potential": 4, "Overall Rating": 8.9,
        "Key Notes": "Spanish-themed project, excellent rental demand (Housing)"
    },
    {
        "Seq": 12, "Project": "Keerthi Regalia", "Area": "Whitefield",
        "Builder": "Keerthi Estates", "Builder Grade": "B+", "Builder Score": 8.4,
        "Price Min (₹/sq.ft)": 9500, "Price Max (₹/sq.ft)": 11000,
        "Nearby IT Hubs": "ITPL, EPIP Zone, GR Tech Park",
        "IT Distance": "4–6 km", "SAP Labs Distance": "5–6 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 8.7,
        "Key Notes": "Good Whitefield connectivity"
    },
    {
        "Seq": 13, "Project": "SUYUG The 1", "Area": "Sarjapur Road",
        "Builder": "Suyug Infra", "Builder Grade": "A-", "Builder Score": 8.8,
        "Price Min (₹/sq.ft)": 10500, "Price Max (₹/sq.ft)": 12000,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ, RMZ Ecoworld",
        "IT Distance": "4–7 km", "SAP Labs Distance": "8–9 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 8.9,
        "Key Notes": "Boutique premium project"
    },
    {
        "Seq": 14, "Project": "Subha Esperanza", "Area": "Sarjapur Road",
        "Builder": "Subha Builders", "Builder Grade": "B+", "Builder Score": 8.3,
        "Price Min (₹/sq.ft)": 8800, "Price Max (₹/sq.ft)": 10200,
        "Nearby IT Hubs": "Wipro SEZ, RGA Tech Park",
        "IT Distance": "5–8 km", "SAP Labs Distance": "8–10 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "5 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 3, "Overall Rating": 8.5,
        "Key Notes": "Budget-friendly project"
    },
    {
        "Seq": 15, "Project": "GR Sitara", "Area": "Whitefield",
        "Builder": "GR Group", "Builder Grade": "B+", "Builder Score": 8.4,
        "Price Min (₹/sq.ft)": 9800, "Price Max (₹/sq.ft)": 11200,
        "Nearby IT Hubs": "ITPL, EPIP Zone",
        "IT Distance": "3–5 km", "SAP Labs Distance": "4–5 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "3 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 8.7,
        "Key Notes": "Strong Whitefield location"
    },
    {
        "Seq": 16, "Project": "Wone 8 Royal Nest", "Area": "Gunjur",
        "Builder": "Wone Developers", "Builder Grade": "B+", "Builder Score": 8.3,
        "Price Min (₹/sq.ft)": 9000, "Price Max (₹/sq.ft)": 10500,
        "Nearby IT Hubs": "RMZ Ecoworld, Embassy Tech Village",
        "IT Distance": "5–7 km", "SAP Labs Distance": "7–8 km",
        "Nearest Metro": "Kadubeesanahalli (Future)", "Metro Distance": "5 km",
        "Society Size": "Small", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 3, "Overall Rating": 8.5,
        "Key Notes": "Boutique community"
    },
    {
        "Seq": 17, "Project": "Mahaveer Ranches", "Area": "Hosa Road",
        "Builder": "Mahaveer Group", "Builder Grade": "A-", "Builder Score": 8.9,
        "Price Min (₹/sq.ft)": 8800, "Price Max (₹/sq.ft)": 10500,
        "Nearby IT Hubs": "Electronic City, RGA Tech Park, Wipro",
        "IT Distance": "4–8 km", "SAP Labs Distance": "10–12 km",
        "Nearest Metro": "Bommasandra", "Metro Distance": "5 km",
        "Society Size": "Large (1,090 units)", "Possession Status": "Ready",
        "Rental Demand": 5, "Resale Potential": 4, "Overall Rating": 8.9,
        "Key Notes": "Large township with good rental demand (JLL Homes)"
    },
    {
        "Seq": 18, "Project": "Mana Dale", "Area": "Kodathi",
        "Builder": "Mana Projects", "Builder Grade": "A-", "Builder Score": 9.0,
        "Price Min (₹/sq.ft)": 10000, "Price Max (₹/sq.ft)": 11800,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ, RMZ Ecoworld",
        "IT Distance": "5–7 km", "SAP Labs Distance": "8–9 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 9.0,
        "Key Notes": "Good boutique project from Mana"
    },
    {
        "Seq": 19, "Project": "Mana Macasa", "Area": "Sarjapur Road",
        "Builder": "Mana Projects", "Builder Grade": "A-", "Builder Score": 9.0,
        "Price Min (₹/sq.ft)": 9800, "Price Max (₹/sq.ft)": 11500,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ",
        "IT Distance": "5–7 km", "SAP Labs Distance": "8–9 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 4, "Resale Potential": 4, "Overall Rating": 8.9,
        "Key Notes": "Good value near IT corridor"
    },
    {
        "Seq": 20, "Project": "DSR Parkway", "Area": "Chikkanayakanahalli",
        "Builder": "DSR Infrastructure", "Builder Grade": "A", "Builder Score": 9.1,
        "Price Min (₹/sq.ft)": 10500, "Price Max (₹/sq.ft)": 12500,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ, RMZ Ecoworld",
        "IT Distance": "4–7 km", "SAP Labs Distance": "8–9 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand": 5, "Resale Potential": 5, "Overall Rating": 9.2,
        "Key Notes": "Strong builder, premium specifications"
    }
]

# Load existing workbook
file_path = "/Users/I771246/Abhi Personal/JavaFullstackNotes/DREAM_HOME/Bangalore_Property_Comparison_Analysis.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Get existing headers from row 1
headers = [cell.value for cell in ws[1]]
print(f"Existing headers: {headers[:19]}")  # Print first 19 relevant headers

# Find the last row with data
last_row = ws.max_row
print(f"Current last row: {last_row}")

# Define styles for consistency
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Add new properties
for prop in new_properties:
    last_row += 1
    for col_idx, header in enumerate(headers, start=1):
        if header is None:
            continue

        cell = ws.cell(row=last_row, column=col_idx)
        cell.value = prop.get(header, "")
        cell.border = thin_border

        # Apply alignment based on column type
        if header in ["Project", "Area", "Builder", "Nearby IT Hubs", "Key Notes"]:
            cell.alignment = left_alignment
        else:
            cell.alignment = center_alignment

        # Bold and highlight for A+ builders
        if header == "Builder Grade" and prop.get("Builder Grade") == "A+":
            cell.font = Font(bold=True)

        # Bold high builder scores
        if header == "Builder Score" and prop.get("Builder Score", 0) >= 9.5:
            cell.font = Font(bold=True)
        elif header == "Builder Score" and prop.get("Builder Score", 0) >= 9.0:
            cell.font = Font(bold=True)

        # Highlight high overall ratings
        if header == "Overall Rating" and prop.get("Overall Rating", 0) >= 9.0:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

# Save the workbook
wb.save(file_path)
print(f"\nSuccessfully added {len(new_properties)} properties to the Excel file!")
print(f"Total rows now: {ws.max_row}")
print(f"New properties added: Seq {new_properties[0]['Seq']} to {new_properties[-1]['Seq']}")
