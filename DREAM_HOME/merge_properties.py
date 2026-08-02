#!/usr/bin/env python3

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Data to merge
new_properties = [
    {
        "Seq": 1, "Project": "Adarsh Tropica", "Area": "Off Sarjapur Road (Gattahalli)",
        "Builder": "Adarsh Developers", "Builder Grade": "A+", "Builder Score": 9.7,
        "Price Min (₹/sq.ft)": 11500, "Price Max (₹/sq.ft)": 13500,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ, RMZ Ecoworld, Embassy Tech Village",
        "IT Distance": "5–8 km", "SAP Labs Distance": "8–10 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "3 km",
        "Society Size": "Large (~700+)", "Possession Status": "Under Construction",
        "Rental Demand (5)": 5, "Resale Potential (5)": 5, "Overall Rating": 9.4,
        "Key Notes": "Excellent builder, strong appreciation corridor"
    },
    {
        "Seq": 2, "Project": "Adarsh Lumina", "Area": "Choodasandra",
        "Builder": "Adarsh Developers", "Builder Grade": "A+", "Builder Score": 9.7,
        "Price Min (₹/sq.ft)": 10800, "Price Max (₹/sq.ft)": 12800,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ, Ecoworld",
        "IT Distance": "5–8 km", "SAP Labs Distance": "8–10 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand (5)": 5, "Resale Potential (5)": 5, "Overall Rating": 9.3,
        "Key Notes": "Premium Adarsh quality, good future appreciation"
    },
    {
        "Seq": 3, "Project": "Poorvi Enchanting", "Area": "Chikkanayakanahalli",
        "Builder": "Poorvi Infinite", "Builder Grade": "A-", "Builder Score": 8.8,
        "Price Min (₹/sq.ft)": 9200, "Price Max (₹/sq.ft)": 10800,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ, Ecoworld",
        "IT Distance": "4–7 km", "SAP Labs Distance": "7–8 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "3 km",
        "Society Size": "Small (99 units)", "Possession Status": "Under Construction",
        "Rental Demand (5)": 4, "Resale Potential (5)": 4, "Overall Rating": 8.9,
        "Key Notes": "Boutique community, low density"
    },
    {
        "Seq": 4, "Project": "Bhavya Pristine", "Area": "Sarjapur Road",
        "Builder": "Bhavya Developers", "Builder Grade": "B+", "Builder Score": 8.3,
        "Price Min (₹/sq.ft)": 8800, "Price Max (₹/sq.ft)": 10000,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ",
        "IT Distance": "6–8 km", "SAP Labs Distance": "9–10 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "4 km",
        "Society Size": "Small", "Possession Status": "Under Construction",
        "Rental Demand (5)": 4, "Resale Potential (5)": 3, "Overall Rating": 8.5,
        "Key Notes": "Budget boutique project"
    },
    {
        "Seq": 5, "Project": "TG Ascent", "Area": "Whitefield",
        "Builder": "TG Developers", "Builder Grade": "B+", "Builder Score": 8.4,
        "Price Min (₹/sq.ft)": 9500, "Price Max (₹/sq.ft)": 11000,
        "Nearby IT Hubs": "ITPL, EPIP Zone, GR Tech Park",
        "IT Distance": "3–5 km", "SAP Labs Distance": "4–5 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "3 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand (5)": 4, "Resale Potential (5)": 4, "Overall Rating": 8.7,
        "Key Notes": "Good Whitefield location"
    },
    {
        "Seq": 6, "Project": "Surbacon Cedar", "Area": "Varthur",
        "Builder": "Surbacon", "Builder Grade": "B+", "Builder Score": 8.4,
        "Price Min (₹/sq.ft)": 8800, "Price Max (₹/sq.ft)": 10000,
        "Nearby IT Hubs": "ITPL, Embassy Tech Village, Ecoworld",
        "IT Distance": "6–8 km", "SAP Labs Distance": "7–8 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "6 km",
        "Society Size": "Small", "Possession Status": "Under Construction",
        "Rental Demand (5)": 4, "Resale Potential (5)": 3, "Overall Rating": 8.5,
        "Key Notes": "Boutique community"
    },
    {
        "Seq": 7, "Project": "SLS Splendor", "Area": "Begur Road",
        "Builder": "SLS Properties", "Builder Grade": "A-", "Builder Score": 8.8,
        "Price Min (₹/sq.ft)": 8500, "Price Max (₹/sq.ft)": 10200,
        "Nearby IT Hubs": "Electronic City Phase 1, Infosys, Wipro",
        "IT Distance": "5–8 km", "SAP Labs Distance": "15+ km",
        "Nearest Metro": "Bommanahalli (Future)", "Metro Distance": "4 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand (5)": 4, "Resale Potential (5)": 4, "Overall Rating": 8.8,
        "Key Notes": "Good value in South Bangalore"
    },
    {
        "Seq": 8, "Project": "SLS Sunflower", "Area": "Electronic City",
        "Builder": "SLS Properties", "Builder Grade": "A-", "Builder Score": 8.8,
        "Price Min (₹/sq.ft)": 8700, "Price Max (₹/sq.ft)": 10300,
        "Nearby IT Hubs": "Infosys, Wipro, Tech Mahindra, TCS",
        "IT Distance": "3–5 km", "SAP Labs Distance": "16+ km",
        "Nearest Metro": "Bommasandra", "Metro Distance": "3 km",
        "Society Size": "Medium", "Possession Status": "Under Construction",
        "Rental Demand (5)": 4, "Resale Potential (5)": 4, "Overall Rating": 8.8,
        "Key Notes": "Strong rental demand"
    },
    {
        "Seq": 9, "Project": "MH Tirumala Heights", "Area": "Whitefield",
        "Builder": "MH Developers", "Builder Grade": "B+", "Builder Score": 8.3,
        "Price Min (₹/sq.ft)": 9000, "Price Max (₹/sq.ft)": 10500,
        "Nearby IT Hubs": "ITPL, EPIP Zone",
        "IT Distance": "4–6 km", "SAP Labs Distance": "5–6 km",
        "Nearest Metro": "Whitefield (Kadugodi)", "Metro Distance": "4 km",
        "Society Size": "Small", "Possession Status": "Under Construction",
        "Rental Demand (5)": 4, "Resale Potential (5)": 3, "Overall Rating": 8.5,
        "Key Notes": "Whitefield boutique project"
    },
    {
        "Seq": 10, "Project": "Emmanuel Heights", "Area": "Sarjapur Road",
        "Builder": "Emmanuel Builders", "Builder Grade": "B", "Builder Score": 8.0,
        "Price Min (₹/sq.ft)": 8200, "Price Max (₹/sq.ft)": 9500,
        "Nearby IT Hubs": "RGA Tech Park, Wipro SEZ",
        "IT Distance": "5–8 km", "SAP Labs Distance": "8–10 km",
        "Nearest Metro": "Carmelaram (Future)", "Metro Distance": "4 km",
        "Society Size": "Small", "Possession Status": "Under Construction",
        "Rental Demand (5)": 3, "Resale Potential (5)": 3, "Overall Rating": 8.2,
        "Key Notes": "Budget option, verify approvals and construction quality"
    }
]

# Load existing workbook
file_path = "/Users/I771246/Abhi Personal/JavaFullstackNotes/DREAM_HOME/Bangalore_Property_Comparison_Analysis.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb.active

# Get existing headers from row 1
headers = [cell.value for cell in ws[1]]
print(f"Existing headers: {headers}")

# Find the last row with data
last_row = ws.max_row
print(f"Current last row: {last_row}")

# Define styles for consistency
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Styling for cells
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_alignment = Alignment(horizontal="center", vertical="center")
left_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# Add new properties
for prop in new_properties:
    last_row += 1
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=last_row, column=col_idx)
        cell.value = prop.get(header, "")
        cell.border = thin_border

        # Apply alignment based on column type
        if header in ["Project", "Area", "Builder", "Nearby IT Hubs", "Key Notes"]:
            cell.alignment = left_alignment
        else:
            cell.alignment = center_alignment

        # Bold and highlight for A+ builders
        if header == "Builder Grade" and prop.get("Builder Grade") == "A+":
            cell.font = Font(bold=True)

        if header == "Builder Score" and prop.get("Builder Score", 0) >= 9.5:
            cell.font = Font(bold=True)

        if header == "Overall Rating" and prop.get("Overall Rating", 0) >= 9.0:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")

# Adjust column widths for better readability
column_widths = {
    'A': 6,   # Seq
    'B': 22,  # Project
    'C': 32,  # Area
    'D': 20,  # Builder
    'E': 12,  # Builder Grade
    'F': 12,  # Builder Score
    'G': 16,  # Price Min
    'H': 16,  # Price Max
    'I': 45,  # Nearby IT Hubs
    'J': 12,  # IT Distance
    'K': 15,  # SAP Labs Distance
    'L': 20,  # Nearest Metro
    'M': 13,  # Metro Distance
    'N': 20,  # Society Size
    'O': 18,  # Possession Status
    'P': 14,  # Rental Demand
    'Q': 16,  # Resale Potential
    'R': 13,  # Overall Rating
    'S': 45   # Key Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Save the workbook
wb.save(file_path)
print(f"\nSuccessfully added {len(new_properties)} properties to the Excel file!")
print(f"Total rows now: {ws.max_row}")
